import csv
import os
import gc
import glob
import logging
from datetime import datetime
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

# Configurar logging para ver progreso en consola y archivo
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[logging.FileHandler(os.path.join(DATA_DIR, 'procesamiento.log')), logging.StreamHandler()]
)

MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
MESES_MAP = {m.lower(): i+1 for i, m in enumerate(MESES)}
DIAS = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', 'Sabado', 'Domingo']


def limpiar_valor(v):
    if v is None:
        return ''
    s = str(v).strip()
    return s


def extraer_streaming_excel(ruta, anio, salida_csv, escribir_encabezado=True):
    """Lee Excel en streaming (bajo consumo) y escribe a CSV"""
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    encabezados = next(rows_iter)
    ncols = len(encabezados)

    modo = 'w' if escribir_encabezado else 'a'
    with open(salida_csv, modo, newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        if escribir_encabezado:
            writer.writerow([
                'carril', 'fecha_trafico', 'fecha', 'hora', 'dia_semana', 'dia_num',
                'mes_num', 'mes', 'anio', 'velocidad_kmh', 'corredor', 'sentido',
                'operacion', 'intensidad_veh_h', 'categoria_1', 'categoria_2',
                'categoria_3', 'categoria_4', 'ocupacion', 'tipo_subsistema',
                'longitud', 'latitud', 'identificador_fv', 'comuna', 'codigo_comuna',
            ])

        filas = 0
        for row in rows_iter:
            row = list(row)
            if not row or all(v is None for v in row):
                continue

            carril = limpiar_valor(row[0]) if ncols > 0 else ''
            fecha_trafico = limpiar_valor(row[1]) if ncols > 1 else ''
            fecha = limpiar_valor(row[2]) if ncols > 2 else ''
            hora = limpiar_valor(row[3]) if ncols > 3 else ''
            dia_semana = limpiar_valor(row[4]) if ncols > 4 else ''
            dia_num = limpiar_valor(row[5]) if ncols > 5 else ''
            mes_nombre = limpiar_valor(row[7]) if ncols > 7 else ''
            velocidad = limpiar_valor(row[9]) if ncols > 9 else ''
            corredor = limpiar_valor(row[10]) if ncols > 10 else ''
            sentido = limpiar_valor(row[11]) if ncols > 11 else ''
            operacion = limpiar_valor(row[12]) if ncols > 12 else ''
            intensidad = limpiar_valor(row[13]) if ncols > 13 else ''
            cat1 = limpiar_valor(row[14]) if ncols > 14 else ''
            cat2 = limpiar_valor(row[15]) if ncols > 15 else ''
            cat3 = limpiar_valor(row[16]) if ncols > 16 else ''
            cat4 = limpiar_valor(row[17]) if ncols > 17 else ''
            ocupacion = limpiar_valor(row[18]) if ncols > 18 else ''
            subsistema = limpiar_valor(row[19]) if ncols > 19 else ''
            longitud = limpiar_valor(row[20]) if ncols > 20 else ''
            latitud = limpiar_valor(row[21]) if ncols > 21 else ''
            id_fv = limpiar_valor(row[22]) if ncols > 22 else ''
            comuna = limpiar_valor(row[23]) if ncols > 23 else ''
            cod_comuna = limpiar_valor(row[24]) if ncols > 24 else ''

            mes_num = MESES_MAP.get(mes_nombre.lower(), 0)

            try:
                hora_int = int(float(hora)) if hora else 0
            except ValueError:
                hora_int = 0

            writer.writerow([
                carril, fecha_trafico, fecha, hora_int, dia_semana, dia_num,
                mes_num, mes_nombre, anio, velocidad, corredor.title(), sentido.upper(),
                operacion.lower(), intensidad, cat1, cat2, cat3, cat4,
                ocupacion, subsistema, longitud, latitud, id_fv, comuna, cod_comuna
            ])
            filas += 1

    wb.close()
    return filas


def main():
    logging.info("=" * 60)
    logging.info("INICIANDO ETL STREAMING - TRÁFICO MEDELLÍN")

    # Detectar carpetas de años automáticamente (carpetas que son números)
    años_disponibles = sorted([int(d) for d in os.listdir(BASE_DIR) if d.isdigit()])

    for anio in años_disponibles:
        logging.info(f"Procesando Año: {anio}")
        base_anio = os.path.join(BASE_DIR, str(anio))
        # Buscar todos los archivos y filtrar manualmente para ser insensibles a mayúsculas
        archivos = sorted([
            os.path.join(base_anio, f) for f in os.listdir(base_anio)
            if f.lower().endswith('.xlsx') and not f.startswith('~$')
        ])
        
        if not archivos:
            logging.warning(f"No se encuentran archivos en la carpeta {anio}")
            continue

        salida_csv = os.path.join(DATA_DIR, f'trafico_{anio}.csv')
        for i, archivo in enumerate(archivos):
            logging.info(f"  Extrayendo: {os.path.basename(archivo)}")
            t0 = datetime.now()
            filas = extraer_streaming_excel(archivo, anio, salida_csv, escribir_encabezado=(i==0))
            delta = (datetime.now() - t0).total_seconds()
            logging.info(f"    Completado: {filas:,} filas en {delta:.1f}s")
            gc.collect()

    logging.info("Procesamiento completado con éxito.")


if __name__ == '__main__':
    main()
