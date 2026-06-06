import pandas as pd
import numpy as np
from sqlalchemy import create_engine, text
import os
import gc
import logging

DB_URL = os.getenv('MYSQL_URL', 'mysql+pymysql://analista:TalentoTech2024!@localhost:3306/trafico_medellin')
engine = create_engine(DB_URL)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[logging.FileHandler(os.path.join(BASE_DIR, 'data', 'mysql_load.log')), logging.StreamHandler()]
)

CHECKPOINT_FILE = os.path.join(BASE_DIR, 'data', 'archivos_procesados.log')

def probar_conexion(engine):
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception as e:
        logging.error(f"No se pudo conectar a MySQL en WSL. ¿Está el servicio activo? Error: {e}")
        return False

with engine.connect() as conn:
    conn.execute(text("""
    CREATE TABLE IF NOT EXISTS trafico (
        id INT AUTO_INCREMENT PRIMARY KEY,
        carril VARCHAR(150), fecha_trafico DATETIME, fecha DATE, hora INT,
        dia_semana VARCHAR(20), dia_num INT, mes_num INT, mes VARCHAR(20),
        anio INT, velocidad_kmh FLOAT, corredor VARCHAR(100),
        sentido VARCHAR(10), operacion VARCHAR(30), intensidad_veh_h FLOAT,
        categoria_1 INT DEFAULT 0, categoria_2 INT DEFAULT 0,
        categoria_3 INT DEFAULT 0, categoria_4 INT DEFAULT 0,
        ocupacion FLOAT, tipo_subsistema VARCHAR(50),
        longitud FLOAT, latitud FLOAT, identificador_fv VARCHAR(10),
        comuna VARCHAR(100), codigo_comuna FLOAT,
        periodo_dia VARCHAR(20), es_hora_pico BOOLEAN DEFAULT FALSE,
        indice_congestion FLOAT DEFAULT 0,
        INDEX idx_anio (anio), INDEX idx_corredor (corredor),
        INDEX idx_hora (hora), INDEX idx_fecha (fecha)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """))
    conn.commit()

def obtener_procesados():
    if not os.path.exists(CHECKPOINT_FILE):
        return set()
    with open(CHECKPOINT_FILE, 'r', encoding='utf-8') as f:
        return set(line.strip() for line in f if line.strip())

def registrar_procesado(filename):
    with open(CHECKPOINT_FILE, 'a', encoding='utf-8') as f:
        f.write(f"{filename}\n")

procesados = obtener_procesados()
if not procesados:
    logging.info("Iniciando carga limpia. Limpiando tabla 'trafico'...")
    with engine.connect() as conn:
        conn.execute(text("TRUNCATE TABLE trafico"))
        conn.commit()
    logging.info("Tabla limpiada con éxito.")
else:
    logging.info(f"Reanudando carga. {len(procesados)} archivos ya están en la base de datos.")

COL_MAP = {
    'Carril': 'carril', 'Fecha Trafico': 'fecha_trafico', 'Fecha': 'fecha',
    'Hora': 'hora', 'dia': 'dia_semana', 'dia-num': 'dia_num',
    'mes-num': 'mes_num', 'mes': 'mes',
    'a├▒o': 'anio', 'año': 'anio',
    'Velocidad (Km/h)': 'velocidad_kmh', 'Corredor': 'corredor',
    'sentido': 'sentido',
    'Operaci├│n': 'operacion', 'Operación': 'operacion',
    'Intensidad (Veh/h)': 'intensidad_veh_h', 'Intensidad': 'intensidad_veh_h',
    'Ocupaci├│n (%)': 'ocupacion', 'Ocupación (%)': 'ocupacion',
    'Ocupaci├│n': 'ocupacion', 'Ocupación': 'ocupacion', 'Ocupacion': 'ocupacion',
    'Tipo de Subsistema': 'tipo_subsistema', 'Longitud': 'longitud',
    'Latitud': 'latitud', 'Identificador (F/V)': 'identificador_fv',
    'Comuna': 'comuna', 'codigo comuna': 'codigo_comuna',
    'Categoria 1 (Veh/h)': 'categoria_1', 'Categoria 2 (Veh/h)': 'categoria_2',
    'Categoria 3 (Veh/h)': 'categoria_3',
    'N├║ vehiculos long 1': 'categoria_1', 'N├║ vehiculos long 2': 'categoria_2',
    'N├║ vehiculos long 3': 'categoria_3', 'N├║ vehiculos long 4': 'categoria_4',
    'N├â┬║ vehiculos long 1': 'categoria_1', 'N├â┬║ vehiculos long 2': 'categoria_2',
    'N├â┬║ vehiculos long 3': 'categoria_3', 'N├â┬║ vehiculos long 4': 'categoria_4',
    'N┬║ vehiculos long 1': 'categoria_1', 'N┬║ vehiculos long 2': 'categoria_2',
    'N┬║ vehiculos long 3': 'categoria_3', 'N┬║ vehiculos long 4': 'categoria_4',
    'Nº vehiculos long 1': 'categoria_1', 'Nº vehiculos long 2': 'categoria_2',
    'Nº vehiculos long 3': 'categoria_3', 'Nº vehiculos long 4': 'categoria_4',
}

DEST_COLS = ['carril', 'fecha_trafico', 'fecha', 'hora', 'dia_semana',
             'dia_num', 'mes_num', 'mes', 'anio', 'velocidad_kmh', 'corredor',
             'sentido', 'operacion', 'intensidad_veh_h',
             'categoria_1', 'categoria_2', 'categoria_3', 'categoria_4',
             'ocupacion',
             'tipo_subsistema', 'longitud', 'latitud', 'identificador_fv',
             'comuna', 'codigo_comuna', 'periodo_dia', 'es_hora_pico']


def limpiar_y_formatear(df, anio):
    """Aplica la lógica de limpieza a un DataFrame (o trozo de él)"""
    df = df.rename(columns=COL_MAP)

    for c in DEST_COLS:
        if c not in df.columns:
            df[c] = np.nan

    # Asegurar que solo tomamos las columnas necesarias antes de procesar
    df = df[[c for c in df.columns if c in DEST_COLS or c in COL_MAP.values()]]

    for col in ['velocidad_kmh', 'intensidad_veh_h', 'ocupacion', 'codigo_comuna', 'longitud', 'latitud']:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    
    df['hora'] = pd.to_numeric(df['hora'], errors='coerce').fillna(0).astype(int)
    df['anio'] = pd.to_numeric(df['anio'], errors='coerce').fillna(anio).astype(int)
    df['dia_num'] = pd.to_numeric(df['dia_num'], errors='coerce').fillna(0).astype(int)
    df['mes_num'] = pd.to_numeric(df['mes_num'], errors='coerce').fillna(0).astype(int)
    df['fecha_trafico'] = pd.to_datetime(df['fecha_trafico'], errors='coerce')
    df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')

    for col in ['carril', 'corredor', 'comuna', 'dia_semana', 'mes']:
        df[col] = df[col].fillna('').astype(str).str.strip().str.title()
    df['sentido'] = df['sentido'].fillna('').astype(str).str.strip().str.upper()
    df['operacion'] = df['operacion'].fillna('').astype(str).str.strip().str.lower()
    df['tipo_subsistema'] = df['tipo_subsistema'].fillna('').astype(str).str.strip()
    df['identificador_fv'] = df['identificador_fv'].fillna('').astype(str).str.strip()

    df['periodo_dia'] = pd.cut(df['hora'], bins=[-1, 5, 11, 17, 23],
                               labels=['Madrugada', 'Mañana', 'Tarde', 'Noche']).astype(str)
    df['es_hora_pico'] = ((df['hora'] >= 6) & (df['hora'] <= 9)) | ((df['hora'] >= 17) & (df['hora'] <= 20))

    return df[DEST_COLS]


def cargar_archivo_robusto(filepath, anio, engine):
    """Carga el archivo usando streaming de filas para evitar errores de memoria"""
    file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
    
    # Si el archivo es muy pesado (> 25MB), entramos en modo streaming manual
    if file_size_mb > 25:
        logging.info(f"    [STREAM] Archivo grande ({file_size_mb:.1f} MB). Procesando por lotes...")
        from openpyxl import load_workbook
        
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        
        headers = [str(c) for c in next(rows)]
        chunk = []
        batch_size = 10000 # Procesamos de 10k en 10k filas
        total_insertadas = 0
        
        for row in rows:
            chunk.append(row)
            if len(chunk) >= batch_size:
                df_chunk = pd.DataFrame(chunk, columns=headers)
                df_chunk = limpiar_y_formatear(df_chunk, anio)
                df_chunk.to_sql('trafico', engine, if_exists='append', index=False, method='multi', chunksize=1000)
                total_insertadas += len(df_chunk)
                chunk = []
                gc.collect()
        
        if chunk: # Último trozo
            df_chunk = pd.DataFrame(chunk, columns=headers)
            df_chunk = limpiar_y_formatear(df_chunk, anio)
            df_chunk.to_sql('trafico', engine, if_exists='append', index=False, method='multi', chunksize=1000)
            total_insertadas += len(df_chunk)
        
        wb.close()
        return total_insertadas
    else:
        # Para archivos pequeños usamos calamine (es mucho más rápido)
        try:
            df = pd.read_excel(filepath, engine='calamine', dtype=str)
        except Exception:
            df = pd.read_excel(filepath, engine='openpyxl', dtype=str)
        
        df_final = limpiar_y_formatear(df, anio)
        df_final.to_sql('trafico', engine, if_exists='append', index=False, method='multi', chunksize=1000)
        return len(df_final)


logging.info("Iniciando carga masiva a MySQL...")
if not probar_conexion(engine):
    exit(1)

años_disponibles = sorted([int(d) for d in os.listdir(BASE_DIR) if d.isdigit()])

for anio in años_disponibles:
    base = os.path.join(BASE_DIR, str(anio))
    # Detección robusta: ignorar temporales (~$) y ser insensible a la extensión
    files = sorted([
        f for f in os.listdir(base) 
        if f.lower().endswith('.xlsx') and not f.startswith('~$')
    ])
    
    if not files:
        logging.warning(f"No se encontraron archivos .xlsx en la carpeta {anio}")
        continue
    
    logging.info(f"Año {anio}: Detectados {len(files)} archivos: {files}")
    for filename in files:
        if filename in procesados:
            continue
            
        f = os.path.join(base, filename)
        try:
            logging.info(f"--- Procesando: {filename} ({anio}) ---")
            filas = cargar_archivo_robusto(f, anio, engine)
            logging.info(f"    [OK] {filas:,} filas insertadas.")
            registrar_procesado(filename)
            gc.collect()
        except MemoryError:
            logging.error(f"    [CRÍTICO] Error de memoria en {filename}. Intentando limpiar...")
            gc.collect()
            continue
        except Exception as e:
            logging.error(f"    [ERROR] No se pudo procesar {filename}: {e}")
            continue

with engine.connect() as conn:
    r = conn.execute(text('SELECT COUNT(*) FROM trafico'))
    print(f'\nTotal: {r.fetchone()[0]:,} reg')
    r = conn.execute(text('SELECT anio, COUNT(*) FROM trafico GROUP BY anio ORDER BY anio'))
    for row in r:
        print(f'  {row[0]}: {row[1]:,}')

# Calcular indice de congestion
logging.info("Calculando índice de congestión (esta fase puede tardar varios minutos)...")
with engine.connect() as conn:
    res = conn.execute(text('SELECT MAX(ocupacion) FROM trafico WHERE ocupacion > 0 AND ocupacion < 1000'))
    max_o = res.fetchone()[0] or 1
    logging.info(f"Valor máximo de ocupación detectado: {max_o}")
    conn.execute(text(f'UPDATE trafico SET indice_congestion = LEAST(ocupacion / {max_o}, 1)'))
    conn.commit()
    logging.info("Índice de congestión calculado.")
    r = conn.execute(text('SELECT AVG(velocidad_kmh), AVG(intensidad_veh_h), AVG(ocupacion) FROM trafico'))
    row = r.fetchone()
    print(f'\nVelocidad media: {row[0]:.1f} km/h')
    print(f'Intensidad media: {row[1]:.0f} veh/h')
    print(f'Ocupación media: {row[2]:.1f}%')
print('OK!')
